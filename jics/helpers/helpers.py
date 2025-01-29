"""
Author: Kamoliddin Usmonov
Date: 2023-04-17
Description: Helper functions for JICS dashboards
"""

import logging

logger = logging.getLogger(__name__)

import datetime
import io
import os
import urllib

import pandas as pd
from openpyxl import load_workbook, utils

import cm_dashboards.ifrs17_accounting.dash_utils as dash_utils
import cm_dashboards.utilities as utilities


def get_df(handler, wvr_path, model_name):
    """
    Get table data from wvr
    :param handler: DB handler
    :param wvr_path: path to wvr file
    :param model_name: model name to be used for getting data
    :return: DataFrame
    """
    df = handler.get_wvr_data(wvr_path, model_name)
    return df


def get_template_df(journal_name, generate_header_rows=True, header=[0]):
    """
    Get DataFrame by reading template csv file
    :param journal_name: journal name (e.g. "AC")
    :return: DataFrame, list of header rows
    """
    current_path = os.path.dirname(os.path.abspath(__file__))
    current_path = os.path.dirname(current_path)
    df = pd.read_csv(
        f"{current_path}\\templates\\{journal_name}.csv",
        header=header,
        encoding="utf-8",
    )
    if len(header) == 1:
        logger.info(f"Renaming columns for template: {journal_name}")
        df.columns = [utilities.remove_substring(col) for col in df.columns]
    header_rows = []
    if generate_header_rows:
        for index, row in df.iterrows():
            first_cell_value = row[0]
            if (
                isinstance(first_cell_value, str)
                and first_cell_value
                and not first_cell_value.startswith(" ")
            ):
                header_rows.append(index)
    return df, header_rows


def convert_to_percentage(value, round_to=2, perform_abs=True):
    """
    Convert value to percentage
    :param value: value to be converted
    :return: value in percentage
    """
    result = value * 100
    # Round to 2 decimal places
    result = round(result, round_to)
    # Convert to positive number
    if perform_abs:
        result = abs(result)
    # Cast to string, add left padding with 0
    result = f"{result:.2f}%"
    return result


def prepare_dropdown_options(df):
    """
    Generate dropdown options
    :param df: DataFrame
    :return: list of dictionaries (value, label, title) for dropdown options
    """
    options = []
    for name in df["Report_Date"]:
        title = name.strftime("Report date %d-%b-%Y")
        option = {
            "label": name,
            "value": name,
            "title": title,
        }
        options.append(option)
    return options


def prepare_table_data(
    df,
    header_rows=[],
    hidden_columns=[],
    filter_column_style=None,
    additional_header=[],
    replace_zero=None,
    multi_index=False,
    show_negative_numbers=True,
):
    """
    Prepare table data
    :param df: DataFrame
    :param header_rows: list of header rows
    :param hidden_columns: list of hidden columns
    :return: tuple (data, columns, conditional_style) for dash_table.DataTable
    """
    # Check if replacement is needed
    if replace_zero is not None:
        df = df.replace({0: replace_zero})

    # Set multiindex table data, column if needed
    if multi_index:
        table_data, columns = dash_utils.set_multi_index_column_names(
            df, show_negative_numbers=show_negative_numbers
        )
    else:
        table_data = df.to_dict("records")
        table_columns = df.columns
        columns = dash_utils.set_column_names(
            table_columns,
            precision=0,
            hidden_columns=hidden_columns,
            show_negative_numbers=show_negative_numbers,
            additional_header=additional_header,
        )
    conditional_style = dash_utils.set_table_style_kics(
        columns, show_negative_numbers=show_negative_numbers
    )
    row_style = dash_utils.set_row_style(header_rows)
    style = conditional_style + row_style

    # Add conditional style for filter column
    if filter_column_style is not None:
        result = dash_utils.set_conditional_style_by_filtering(
            column=filter_column_style
        )
        style = style + result

    return table_data, columns, style


def get_table_name_by_journal_code(journal_code, journal_type="ordinary"):
    """
    Get table name by journal type
    :param journal_code: journal code (e.g. "1-1")
    :param journal_type: journal type (e.g. "ordinary", "with_selection", "mixed")
    :return: table name (str) or None
    """
    tables = {
        "ordinary": {
            "tab_1": "A_ESR",
            "tab_2": "A_Balance_Sheet",
            "tab_3": "A_MOCE",
            "tab_4": "A_MOCE_RC_Future",
            "tab_5": "A_MOCE_Run_Off",
            "tab_6": "A_RC",
            "tab_7": "A_Life_Risk",
            "tab_8": "A_Non_Life_Risk",
            "tab_9": "A_Cat_Risk",
            "tab_10": "A_Market_Risk",
            "tab_11": "A_Credit_Risk",
            "tab_12": "A_Op_Risk",
            "tab_13": "A_Non_Insurance",
        },
        "with_selection": None,
        "mixed": None,
    }

    result = None
    try:
        result = tables[journal_type][journal_code]
    except KeyError:
        pass
    logger.info(
        f"Selected journal: {journal_code} - {journal_type}, table name: {result}"
    )
    return result


def apply_formatting(cell_value, lookup_dict, perform_abs=True):
    """
    Apply formatting to cell value
    :param cell_value: cell value
    :param lookup_dict: lookup dictionary
    :return: formatted cell value with replacement value if applicable
    """
    if (
        pd.isna(cell_value)
        or pd.isnull(cell_value)
        or cell_value == ""
        or isinstance(cell_value, int)
    ):
        if str(cell_value).lower() == "true":
            return str(cell_value).upper()
        return cell_value

    percentage_variables = get_percentage_variables()
    # Remove substring from cell value if possible
    try:
        new_value = lookup_dict.get(cell_value, cell_value)
    except Exception as e:
        logger.info("Error occurred: {}".format(e))
        new_value = cell_value

    # Cast cell value to float if possible
    if not isinstance(new_value, (int, float, datetime.date)) and new_value.isdigit():
        new_value = float(new_value)

    # Check if cell value is percentage
    if cell_value in percentage_variables:
        new_value = convert_to_percentage(new_value, perform_abs=perform_abs)
    elif (
        cell_value.startswith("Run_Off_")
        or cell_value.startswith("Param1_")
        or cell_value.startswith("Param2_")
        or cell_value.startswith("Stress_Factor_")
    ):
        new_value = convert_to_percentage(new_value, perform_abs=perform_abs)
    # logger.info(f"Cell value: {cell_value}, new value: {new_value}")
    return new_value


def replace_template_values(
    template_df,
    lookup_dict,
    perform_abs=True,
):
    """
    Replace template values by checking
    """
    df = template_df.copy()

    # Use applymap to replace values if there is no formula in the template
    df = df.applymap(
        lambda x: apply_formatting(x, lookup_dict, perform_abs=perform_abs)
    )
    return df


def get_percentage_variables():
    """
    Retrive percentage variables
    :return: list of percentage variables (list)
    """
    results = [
        "ESR",
        "Effective_Tax_Rate",
        "Tier2_DED_from_RB_Factor",
        "Tier2_DED_from_Soft_Factor",
        "Ins_RC_x_20_Percent_Factor",
        "a_Ins_RC_Calc_Factor",
        "b_Ins_RC_Calc_Factor",
        "d_Ins_RC_Calc_Factor",
        "Effective_Tax_Rate",
        "RC_Life_Percentile_Factor",
        "RC_NonLife_Percentile_Factor",
        "Effective_Tax_Rate",
        "DiscretionaryBenRatio_MR_CUR1",
        "DiscretionaryBenRatio_MR_CUR2",
        "DiscretionaryBenRatio_MR_CUR3",
        "DiscretionaryBenRatio_MR_CUR4",
        "DiscretionaryBenRatio_MR_CUR5",
        "DiscretionaryBenRatio_MR_CUR6",
        "DiscretionaryBenRatio_MR_CUR7",
        "DiscretionaryBenRatio_LU_CUR1",
        "DiscretionaryBenRatio_LU_CUR2",
        "DiscretionaryBenRatio_LU_CUR3",
        "DiscretionaryBenRatio_LU_CUR4",
        "DiscretionaryBenRatio_LU_CUR5",
        "DiscretionaryBenRatio_LU_CUR6",
        "DiscretionaryBenRatio_LU_CUR7",
        "DiscretionaryBenRatio_LU",
        "DiscretionaryBenRatio_MR",
        "DiscretionaryBenRatio_LD",
        "DiscretionaryBenRatio_LD_CUR1",
        "DiscretionaryBenRatio_LD_CUR2",
        "DiscretionaryBenRatio_LD_CUR3",
        "DiscretionaryBenRatio_LD_CUR4",
        "DiscretionaryBenRatio_LD_CUR5",
        "DiscretionaryBenRatio_LD_CUR6",
        "DiscretionaryBenRatio_LD_CUR7",
        "DiscretionaryBenRatio_TUD_CUR1",
        "DiscretionaryBenRatio_TUD_CUR2",
        "DiscretionaryBenRatio_TUD_CUR3",
        "DiscretionaryBenRatio_TUD_CUR4",
        "DiscretionaryBenRatio_TUD_CUR5",
        "DiscretionaryBenRatio_TUD_CUR6",
        "DiscretionaryBenRatio_TUD_CUR7",
        "DiscretionaryBenRatio_TUD",
        "DiscretionaryBenRatio_TDU_CUR1",
        "DiscretionaryBenRatio_TDU_CUR2",
        "DiscretionaryBenRatio_TDU_CUR3",
        "DiscretionaryBenRatio_TDU_CUR4",
        "DiscretionaryBenRatio_TDU_CUR5",
        "DiscretionaryBenRatio_TDU_CUR6",
        "DiscretionaryBenRatio_TDU_CUR7",
        "DiscretionaryBenRatio_TDU",
        "DiscretionaryBenRatio_Spr_Up",
        "DiscretionaryBenRatio_Spr_Down",
        "DiscretionaryBenRatio_DM",
        "DiscretionaryBenRatio_EM",
        "DiscretionaryBenRatio_HP",
        "DiscretionaryBenRatio_OTHE",
        "DiscretionaryBenRatio_Vol",
        "Equity_Credit_Risk_Factor",
        "Conc_excl_RE_bef_Limit_Factor",
        "Conc_Threshold_RE",
        "Conc_Real_Estate_Factor",
        "Opr_P_NL_Factor",
        "Opr_Growth_NL_Factor",
        "Opr_Liab_NL_Factor",
        "Opr_P_L_Risk_Factor",
        "Opr_Growth_L_Risk_Factor",
        "Opr_Liab_L_Risk_Factor",
        "Opr_Liab_L_NonRisk_Factor",
        "Opr_L_Risk_Limit_Factor",
        "Opr_Threshold",
        "Cost_of_Capital_ESR",
        "Tier_1_limited_Factor",
        "Tier_1_limited_PLAM_Factor",
        "Tier_2_limited_Factor",
        "Tier_1_Limited_Plus_2_Factor",
        "Tier_2_basket_Factor",
        "Tier_2_Non_Paid_Up_Factor",
        "Corr_Mkt_ESR_1_1",
        "Corr_Mkt_ESR_1_2",
        "Corr_Mkt_ESR_1_3",
        "Corr_Mkt_ESR_1_4",
        "Corr_Mkt_ESR_1_5",
        "Corr_Mkt_ESR_1_6",
        "Corr_Mkt_ESR_1_7",
        "Corr_Mkt_ESR_2_1",
        "Corr_Mkt_ESR_2_2",
        "Corr_Mkt_ESR_2_3",
        "Corr_Mkt_ESR_2_4",
        "Corr_Mkt_ESR_2_5",
        "Corr_Mkt_ESR_2_6",
        "Corr_Mkt_ESR_2_7",
        "Corr_Mkt_ESR_3_1",
        "Corr_Mkt_ESR_3_2",
        "Corr_Mkt_ESR_3_3",
        "Corr_Mkt_ESR_3_4",
        "Corr_Mkt_ESR_3_5",
        "Corr_Mkt_ESR_3_6",
        "Corr_Mkt_ESR_3_7",
        "Corr_Mkt_ESR_4_1",
        "Corr_Mkt_ESR_4_2",
        "Corr_Mkt_ESR_4_3",
        "Corr_Mkt_ESR_4_4",
        "Corr_Mkt_ESR_4_5",
        "Corr_Mkt_ESR_4_6",
        "Corr_Mkt_ESR_4_7",
        "Corr_Mkt_ESR_5_1",
        "Corr_Mkt_ESR_5_2",
        "Corr_Mkt_ESR_5_3",
        "Corr_Mkt_ESR_5_4",
        "Corr_Mkt_ESR_5_5",
        "Corr_Mkt_ESR_5_6",
        "Corr_Mkt_ESR_5_7",
        "Corr_Mkt_ESR_6_1",
        "Corr_Mkt_ESR_6_2",
        "Corr_Mkt_ESR_6_3",
        "Corr_Mkt_ESR_6_4",
        "Corr_Mkt_ESR_6_5",
        "Corr_Mkt_ESR_6_6",
        "Corr_Mkt_ESR_6_7",
        "Corr_Mkt_ESR_7_1",
        "Corr_Mkt_ESR_7_2",
        "Corr_Mkt_ESR_7_3",
        "Corr_Mkt_ESR_7_4",
        "Corr_Mkt_ESR_7_5",
        "Corr_Mkt_ESR_7_6",
        "Corr_Mkt_ESR_7_7",
        "Corr_Equity_ESR_1_1",
        "Corr_Equity_ESR_1_2",
        "Corr_Equity_ESR_1_3",
        "Corr_Equity_ESR_1_4",
        "Corr_Equity_ESR_2_1",
        "Corr_Equity_ESR_2_2",
        "Corr_Equity_ESR_2_3",
        "Corr_Equity_ESR_2_4",
        "Corr_Equity_ESR_3_1",
        "Corr_Equity_ESR_3_2",
        "Corr_Equity_ESR_3_3",
        "Corr_Equity_ESR_3_4",
        "Corr_Equity_ESR_4_1",
        "Corr_Equity_ESR_4_2",
        "Corr_Equity_ESR_4_3",
        "Corr_Equity_ESR_4_4",
        "Ins_RC_Calc_Factor",
        "RC_Liability_Like_Risk_Factor",
        "RC_Motor_Like_Risk_Factor",
        "RC_Property_Like_Risk_Factor",
        "RC_Other_Risk_Factor",
        "Corr_RC_ESR_1_1",
        "Corr_RC_ESR_1_2",
        "Corr_RC_ESR_1_3",
        "Corr_RC_ESR_1_4",
        "Corr_RC_ESR_1_5",
        "Corr_RC_ESR_2_1",
        "Corr_RC_ESR_2_2",
        "Corr_RC_ESR_2_3",
        "Corr_RC_ESR_2_4",
        "Corr_RC_ESR_2_5",
        "Corr_RC_ESR_3_1",
        "Corr_RC_ESR_3_2",
        "Corr_RC_ESR_3_3",
        "Corr_RC_ESR_3_4",
        "Corr_RC_ESR_3_5",
        "Corr_RC_ESR_4_1",
        "Corr_RC_ESR_4_2",
        "Corr_RC_ESR_4_3",
        "Corr_RC_ESR_4_4",
        "Corr_RC_ESR_4_5",
        "Corr_RC_ESR_5_1",
        "Corr_RC_ESR_5_2",
        "Corr_RC_ESR_5_3",
        "Corr_RC_ESR_5_4",
        "Corr_RC_ESR_5_5",
        "MA_Efffect_Ratio_RC",
        "MA_Effect_Ratio_RC_Life",
        "MA_Effect_Ratio_RC_CAT",
        "MA_Effect_Ratio_RC_Market",
        "MA_Effect_Ratio_RC_Credit",
        "MA_Effect_Ratio_Sum_of_Risks",
        "Corr_Life_ESR_1_1",
        "Corr_Life_ESR_1_2",
        "Corr_Life_ESR_1_3",
        "Corr_Life_ESR_1_4",
        "Corr_Life_ESR_1_5",
        "Corr_Life_ESR_2_1",
        "Corr_Life_ESR_2_2",
        "Corr_Life_ESR_2_3",
        "Corr_Life_ESR_2_4",
        "Corr_Life_ESR_2_5",
        "Corr_Life_ESR_3_1",
        "Corr_Life_ESR_3_2",
        "Corr_Life_ESR_3_3",
        "Corr_Life_ESR_3_4",
        "Corr_Life_ESR_3_5",
        "Corr_Life_ESR_4_1",
        "Corr_Life_ESR_4_2",
        "Corr_Life_ESR_4_3",
        "Corr_Life_ESR_4_4",
        "Corr_Life_ESR_4_5",
        "Corr_Life_ESR_5_1",
        "Corr_Life_ESR_5_2",
        "Corr_Life_ESR_5_3",
        "Corr_Life_ESR_5_4",
        "Corr_Life_ESR_5_5",
        "NL_Divers_P_Res_Risk_Factor",
        "NL_Liability_Like_Risk_Factor",
        "NL_Motor_Like_Risk_Factor",
        "NL_Property_Like_Risk_Factor",
        "NL_Other_Risk_Factor",
        "NL_Divers_Categories_Factor",
        "NL_Risk_Geog_Divers_Factor",
        "Corr_DM_ESR",
        "Corr_EM_ESR",
        "Conc_001_Limit",
        "Conc_002_Limit",
        "Conc_003_Limit",
        "Conc_004_Limit",
        "Conc_005_Limit",
        "Conc_006_Limit",
        "Conc_007_Limit",
        "Conc_008_Limit",
        "Conc_009_Limit",
        "Conc_010_Limit",
        "Conc_011_Limit",
        "Conc_012_Limit",
        "Conc_013_Limit",
        "Conc_014_Limit",
        "Conc_015_Limit",
        "Conc_016_Limit",
        "Conc_017_Limit",
        "Conc_018_Limit",
        "Conc_019_Limit",
        "Conc_020_Limit",
        "Conc_021_Limit",
        "Conc_022_Limit",
        "Conc_023_Limit",
        "Conc_024_Limit",
        "Conc_025_Limit",
        "Conc_026_Limit",
        "Conc_027_Limit",
        "Conc_028_Limit",
        "Conc_029_Limit",
        "Conc_030_Limit",
        "Conc_031_Limit",
        "Conc_032_Limit",
        "Conc_033_Limit",
        "Conc_034_Limit",
        "Conc_035_Limit",
        "Conc_036_Limit",
        "Conc_037_Limit",
        "Conc_038_Limit",
        "Conc_039_Limit",
        "Conc_040_Limit",
        "Conc_041_Limit",
        "Conc_042_Limit",
        "Conc_043_Limit",
        "Conc_044_Limit",
        "Conc_045_Limit",
        "Conc_046_Limit",
        "Conc_047_Limit",
        "Conc_048_Limit",
        "Conc_049_Limit",
        "Conc_050_Limit",
        "Conc_051_Limit",
        "Conc_052_Limit",
        "Conc_053_Limit",
        "Conc_054_Limit",
        "Conc_055_Limit",
        "Conc_056_Limit",
        "Conc_057_Limit",
        "Conc_058_Limit",
        "Conc_059_Limit",
        "Conc_060_Limit",
        "Conc_061_Limit",
        "Conc_062_Limit",
        "Conc_063_Limit",
        "Conc_064_Limit",
        "Conc_065_Limit",
        "Conc_066_Limit",
        "Conc_067_Limit",
        "Conc_068_Limit",
        "Conc_069_Limit",
        "Conc_070_Limit",
        "Conc_071_Limit",
        "Conc_072_Limit",
        "Conc_073_Limit",
        "Conc_074_Limit",
        "Conc_075_Limit",
        "Conc_076_Limit",
        "Conc_077_Limit",
        "Conc_078_Limit",
        "Conc_079_Limit",
        "Conc_080_Limit",
        "Conc_081_Limit",
        "Conc_082_Limit",
        "Conc_083_Limit",
        "Conc_084_Limit",
        "Conc_085_Limit",
        "Conc_086_Limit",
        "Conc_087_Limit",
        "Conc_088_Limit",
        "Conc_089_Limit",
        "Conc_090_Limit",
        "Conc_091_Limit",
        "Conc_092_Limit",
        "Conc_093_Limit",
        "Conc_094_Limit",
        "Conc_095_Limit",
        "Conc_096_Limit",
        "Conc_097_Limit",
        "Conc_098_Limit",
        "Conc_099_Limit",
        "Conc_100_Limit",
        "RC_Currency_Correlation_Factor",
    ]
    return results


def replace_header_values(df, output_dict, numeric_format=False):
    """
    Replace header values
    :param df: DataFrame with update needed values
    :param output_dict: Dictionary with actual values
    :return: DataFrame with updated values
    """
    headers = df.columns.to_list()
    replace_dict = dict()
    percentage_variables = get_percentage_variables()
    for header in headers:
        if isinstance(header, tuple):
            for column in header:
                # logger.info(f"Checking column: {column}")
                # Check if output exists for column, and it's a percentage variable
                if column in percentage_variables and column in output_dict.keys():
                    result = convert_to_percentage(output_dict[column])
                    # logger.info(f"Found percentage variable {column}: {result}")
                    replace_dict[column] = result
                    continue

                # Check if column is in output_dict
                if column in output_dict.keys():
                    result = output_dict[column]
                    if numeric_format:
                        result = "{:,.0f}".format(result)
                    replace_dict[column] = result
        else:
            # Check if output exists for column, and it's a percentage variable
            if header in percentage_variables and header in output_dict.keys():
                result = convert_to_percentage(output_dict[header])
                # logger.info(f"Found percentage variable {header}: {result}")
                replace_dict[header] = str(result)
                continue
            # Check if column is in output_dict
            if header in output_dict.keys():
                result = output_dict[header]
                if numeric_format:
                    result = "{:,.0f}".format(result)
                replace_dict[header] = result
    # logger.info(f"Replace dict: {replace_dict}")
    df.rename(columns=replace_dict, inplace=True)
    return df


def create_url_query_string(**params):
    """
    Create url query string from given params
    :param params: list of params
    :return: url query string
    """
    query_string = urllib.parse.urlencode({key: value for key, value in params.items()})
    logger.info("Query string: {}".format(query_string))
    return query_string


def convert_query_string_to_dict(query_string):
    """
    Convert query string to dictionary
    :param query_string: query string
    :return: dictionary
    """
    return dict(urllib.parse.parse_qsl(query_string))


def generate_excel(results):
    """
    This function is used to generate customized excel file from dataframes
    :param dataframes: list of dataframes
    :return: excel file (bytes)
    """
    logger.info("Writing dataframes to excel...")
    # Create a Pandas Excel writer object
    output = io.BytesIO()
    writer = pd.ExcelWriter(
        output,
        engine="xlsxwriter",
        engine_kwargs={"options": {"nan_inf_to_errors": True}},
    )
    logger.info("Total sheets: {}".format(len(results)))
    for sheet_name, result in results.items():
        config = result.get("config", {})
        dataframes = result.get("result", [])
        logger.info(
            "Sheet name: {}, total dashboards: {}".format(sheet_name, len(dataframes))
        )
        # Create a workbook and worksheet objects
        workbook = writer.book
        worksheet = workbook.add_worksheet(sheet_name)

        # Define header and cell formats for coloring the dataframes
        header_format = workbook.add_format(
            {
                "bg_color": "#D3D3D3",
                "align": "center",
                "valign": "vcenter",
                "border": 1,
                "bold": True,
                "text_wrap": True,
                "font_name": "Arial",
            }
        )
        first_cell_format = workbook.add_format(
            {
                "bg_color": "#D3D3D3",
                "align": "left",
                "valign": "vcenter",
                "text_wrap": False,
                "border": 1,
                "font_name": "Arial",
            }
        )
        second_cell_format = workbook.add_format(
            {
                "bg_color": "#D3D3D3",
                "align": "center",
                "valign": "vcenter",
                "font_name": "Arial",
                "border": 1,
            }
        )
        blank_cell_format = workbook.add_format(
            {
                "bg_color": "#D3D3D3",
                "pattern": 4,
                "border": 1,
            }
        )
        zero_cell_format = workbook.add_format(
            {
                "bg_color": "#DDEBF7",
                "border": 0,
            }
        )
        default_cell_format = workbook.add_format(
            {
                "bg_color": "#FFFAA0",
                "align": "right",
                "valign": "vcenter",
                "font_name": "Arial",
                "border": 1,
                "num_format": "#,##0.00",
            }
        )

        # Set the row and column offsets for writing dataframes
        row_offset = 0
        col_offset = 0

        # Loop through each dataframe and write its data and headers to the sheet
        for df in dataframes:
            # Write the headers with the corresponding formatting
            if isinstance(df.columns, pd.MultiIndex):
                nlevels = len(df.columns.levels)
            else:
                nlevels = 1
            logger.info(f"DataFrame levels: {nlevels}")

            # Keep track of the last column for each level
            previous_column = None
            previous_column_per_level = {level: None for level in range(nlevels)}
            duplicate_headers = {level: dict() for level in range(nlevels)}

            # Loop through each column in the dataframe and write its header
            # logger.info(f"Last seen column: {previous_column_per_level}")
            # logger.info(f"Duplicate headers: {duplicate_headers}")
            for col_num, value in enumerate(df.columns.values, start=col_offset):
                # Check if column is multi-indexed
                if isinstance(value, tuple):
                    # logger.info(f"{col_num} Writing multi-index header: {value}")
                    for i, org_column_name in enumerate(value):
                        column_name = utilities.remove_substring(org_column_name)
                        # logger.info(f"({i}) Writing sub-header: {column_name}")

                        # Write the header to the sheet
                        worksheet.write(
                            row_offset + i,
                            col_offset + col_num,
                            column_name,
                            header_format,
                        )
                        # logger.info(f"Last column: {previous_column_per_level}")

                        if previous_column_per_level[i] != column_name:
                            # Keep track of the last column for current level
                            previous_column_per_level[i] = column_name
                            continue

                        # logger.info(f"Found repeating header: {column_name}")
                        # logger.info(f"Duplicate headers: {duplicate_headers}")
                        # Update the duplicate header details or add new one
                        if column_name in duplicate_headers[i]:
                            # logger.info(
                            #     f"Updating repeating empty header: {column_name}"
                            # )
                            header_details = duplicate_headers[i][column_name]
                            if header_details["end_col"] == col_num - 1:
                                header_details["count"] += 1
                                header_details["end_col"] = col_num
                            else:
                                logger.info(
                                    "Skipping repeating empty header at the same level to avoid merge conflict: {}".format(
                                        column_name
                                    )
                                )
                        else:
                            details = {
                                "count": 2,
                                "start_col": col_num - 1,
                                "end_col": col_num,
                                "row_index": row_offset + i,
                            }
                            duplicate_headers[i][column_name] = details

                        # Keep track of the last column for current level
                        previous_column_per_level[i] = column_name
                else:
                    # logger.info("Writing ordinary header: {}".format(value))
                    column_name = utilities.remove_substring(value)
                    worksheet.write(
                        row_offset,
                        col_offset + col_num,
                        column_name,
                        header_format,
                    )

                    # Check if the header is repeating
                    if column_name != "#" and column_name == previous_column:
                        if column_name in duplicate_headers:
                            duplicate_headers[i][column_name]["count"] += 1
                            duplicate_headers[i][column_name]["end_col"] = col_num
                        else:
                            duplicate_headers[i][column_name] = {
                                "count": 2,
                                "start_col": col_num - 1,
                                "end_col": col_num,
                                "row_index": row_offset,
                            }

                    # Set the previous column to the current column
                    previous_column = column_name

            # Merge headers horizontally if any are repeating
            # logger.info("Duplicate sub-values: {}".format(duplicate_headers))
            for level, columns_info in duplicate_headers.items():
                if columns_info is None:
                    continue
                for column_name, info in columns_info.items():
                    # logger.info(f"Level ({level}) - merging sub-value: {column_name}")
                    worksheet.merge_range(
                        info["row_index"],
                        info["start_col"],
                        info["row_index"],
                        info["end_col"],
                        column_name,
                        header_format,
                    )

            # Adjust the row offset for writing data
            nlevels -= 1

            # Write the data with the corresponding formatting (with merged cells)
            for j in range(len(df.columns)):
                # Initialize variables for tracking consecutive empty cells
                # logger.info("Writing column index: {}".format(j))
                start_empty_row = None
                empty_rows = 0

                for i in range(len(df)):
                    # Get the value at the current cell
                    value = df.iloc[i, j]
                    # logger.info("- Writing value: {}: {} - {}".format(value, i, j))

                    # Adjust formatting for non-empty cells
                    if j == 0:
                        cell_format = first_cell_format
                    elif j == 1:
                        cell_format = second_cell_format
                    else:
                        cell_format = default_cell_format

                    # Do not merge first and second column cells
                    if j == 0 or j == 1:
                        worksheet.write(
                            row_offset + i + nlevels + 1,
                            col_offset + j,
                            value,
                            cell_format,
                        )
                        continue

                    # Check if the cell is empty
                    if pd.isna(value) or value == "":
                        empty_rows += 1
                        # If this is the first empty cell in a consecutive series, record the row index
                        if start_empty_row is None:
                            start_empty_row = i

                    else:
                        # If there were consecutive empty cells, merge them
                        if empty_rows > 1:
                            # logger.info("Inside loop")
                            # logger.info("Merging empty cells: {}".format(empty_rows))
                            # logger.info(
                            #     "Start row: {}, End row: {}".format(
                            #         row_offset + start_empty_row + nlevel,
                            #         row_offset + i - 1 + nlevel,
                            #     )
                            # )
                            worksheet.merge_range(
                                row_offset + start_empty_row + nlevels + 1,
                                col_offset + j,
                                row_offset + i + nlevels,
                                col_offset + j,
                                "",
                                blank_cell_format,
                            )
                        empty_rows = 0
                        start_empty_row = None

                        # Write the value to the cell with the corresponding formatting
                        # logger.info("Writing non-empty value: {}".format(value))
                        # logger.info(
                        #     "Start row: {}".format(row_offset + i + nlevel - empty_rows)
                        # )
                        worksheet.write(
                            row_offset + i + nlevels - empty_rows + 1,
                            col_offset + j,
                            value,
                            cell_format,
                        )

                # If there were consecutive empty cells at the end of the column, merge them
                if empty_rows > 1:
                    # logger.info("After loop")
                    # logger.info("Merging empty cells: {}".format(empty_rows))
                    # logger.info(
                    #     "Start row: {}, End row: {}".format(
                    #         row_offset + start_empty_row + nlevel,
                    #         row_offset + i + nlevel,
                    #     )
                    # )
                    worksheet.merge_range(
                        row_offset + start_empty_row + nlevels + 1,
                        col_offset + j,
                        row_offset + i + nlevels + 1,
                        col_offset + j,
                        "",
                        blank_cell_format,
                    )

            # Apply formating for empty cells
            if len(df) > 0:
                worksheet.conditional_format(
                    row_offset + nlevels + 1,
                    col_offset + 1,
                    row_offset + nlevels + len(df.index),
                    col_offset + len(df.columns) - 1,
                    {
                        "type": "blanks",
                        "format": blank_cell_format,
                    },
                )

                # Apply formating for zero valued cells
                worksheet.conditional_format(
                    row_offset + nlevels + 1,
                    col_offset + 1,
                    row_offset + nlevels + len(df.index),
                    col_offset + len(df.columns) - 1,
                    {
                        "type": "cell",
                        "criteria": "==",
                        "value": 0,
                        "format": zero_cell_format,
                    },
                )

            # Update the row offset for the next dataframe
            # Add a gap of 1 rows between dataframes
            # logger.info(f"Current row offset: {row_offset}")
            row_offset += len(df) + nlevels + 2
            # logger.info(f"Next row offset: {row_offset}")

        sheet_color = config.get("sheet_color", "red")
        sheet_with = config.get("sheet_width", 25)
        logger.info(f"Setting sheet color: {sheet_color}, width: {sheet_with}")

        # Write the border from beginning to end of the current worksheet
        border_data = ["#"] * (row_offset - 1)
        worksheet.write_column(
            0,
            sheet_with,
            border_data,
            header_format,
        )

        # Update column width (A, B, C-J) and set tab color in the current worksheet
        worksheet.set_column(0, 0, 80)
        worksheet.set_column(1, 1, 20)
        worksheet.set_column(2, sheet_with - 2, 30)
        worksheet.set_column(sheet_with - 1, sheet_with, 4)
        worksheet.set_tab_color(sheet_color)
        # Set zoom level and protect the worksheet from editing
        worksheet.set_zoom(85)
        # worksheet.protect("JICS")

    # Save the Excel file
    writer.save()

    # Close the Pandas Excel writer and workbook objects
    writer.close()

    # Reset the buffer position and send the Excel file to the user
    output.seek(0)

    return output


def generate_report_file(results):
    """
    Write data to the template excel file (openpyxl engine)
    :param results: Results (dict) - sheet name, its data and dashboard ids
    :return updated excel file (in memory buffer)
    """
    current_path = os.path.dirname(os.path.abspath(__file__))
    current_path = os.path.dirname(current_path)
    template_file = os.path.join(current_path, "templates", "JICS_template.xlsx")
    book = load_workbook(template_file, data_only=True)
    book.calculation_on_load = True
    book.active = 0
    book.active.cell(row=1, column=1)

    # Loop through each sheet that contains its data and dashboard ids
    black_list = ["HEADER"]
    try:
        for sheet_name in results.keys():
            logger.info(f"Processing sheet name: {sheet_name}")
            sheet_result = results[sheet_name]["data"]
            dashboard_ids = results[sheet_name]["dashboard_ids"]
            for dashboard_id in dashboard_ids:
                if dashboard_id in black_list:
                    logger.info(f"Skipping dashboard id: {dashboard_id}")
                    continue

                current_range = f"{dashboard_id}_block"
                current_range = current_range.replace("-", "_")
                try:
                    coords, sheet = utilities.get_range_destination(book, current_range)
                except Exception:
                    continue

                boundaries = utils.cell.range_boundaries(coords)
                first_col = boundaries[0]
                first_row = boundaries[1]
                last_col = boundaries[2]
                last_row = boundaries[3]

                for row in sheet.iter_rows(
                    min_row=first_row,
                    max_row=last_row,
                    min_col=first_col,
                    max_col=last_col,
                ):
                    for cell in row:
                        if cell.value is None or cell.value == "":
                            continue

                        if cell.value in sheet_result:
                            cell.value = sheet_result[cell.value]

    except Exception as error:
        logger.info(f"Error while processing sheet: {sheet_name}")
        logger.info(f"Error: {error}")
        raise utilities.add_exception_info(error, "Error while processing report file")

    # Save the changes to the memory buffer and close the workbook
    in_mem_file = io.BytesIO()
    book.save(in_mem_file)
    book.close()

    # Reset the pointer to the beginning of the file
    in_mem_file.seek(0)
    return in_mem_file
