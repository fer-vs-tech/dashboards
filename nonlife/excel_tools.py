from openpyxl import load_workbook, utils


def load_workbook_range(range_string, worksheet):
    """
    Loop through range data and add to list
    """
    data_rows = []
    for row in worksheet[range_string]:
        data_rows.append([cell.value for cell in row])
    return data_rows


def get_named_ranges(book):
    """
    Get list of named ranges in the Excel workbook,
    filtering out non-assumption ones
    """
    named_ranges = list()
    # Get list of defined names
    for r in book.defined_names.definedName:
        # Filter out invalid and non-user entries
        if (
            not (r.hidden)
            and not (r.attr_text.startswith("["))
            and r.attr_text.find("#REF!") == -1
        ):
            named_ranges.append(r.name)
    return named_ranges


def get_range_destination(book, range_name):
    """
    Get coordinates and sheet of this range
    """
    # Get location information of this range
    dest = list(book.defined_names[range_name].destinations)
    # Excel cell coordinates of the range
    coords = dest[0][1].replace("$", "")
    # Get the sheet object
    sheet = book[dest[0][0]]
    return coords, sheet


def find_assumption_dimensions(data_list, range_string, worksheet):
    """
    Find the number of dimensions in an assumption table and name of first dimension
    """
    # Get boundaries of named range
    boundaries = utils.cell.range_boundaries(range_string)
    # Figure out coordinates
    first_cell_col_letter = utils.cell.get_column_letter(boundaries[0])
    first_cell_rownum = boundaries[1]
    last_cell_col_letter = utils.cell.get_column_letter(boundaries[2])
    # Row above the named range
    prior_rownum = first_cell_rownum - 1
    # Build address of row above the named range to find the first dimension name and position
    prior_row_range_string = "{0}{1}:{2}{3}".format(
        first_cell_col_letter, prior_rownum, last_cell_col_letter, prior_rownum
    )
    dimension_count = 0
    first_dimension_name = None

    # Search for first dimension name in above row
    for cell in worksheet[prior_row_range_string][0]:
        dimension_count += 1
        if dimension_count > 5:
            # No dimension title found
            break
        if cell.value:
            first_dimension_name = cell.value
            break
    if first_dimension_name is None:
        # No dimension title found, assume single dimension
        dimension_count = 1

        # for single dimension, get first dimension name from first named range header
        firstcell = worksheet[first_cell_col_letter + str(first_cell_rownum)]
        if firstcell.value:
            first_dimension_name = firstcell.value
        else:
            # If empty, set to Category
            first_dimension_name = "Category"

    dimension_list = get_dimension_names(
        data_list, first_dimension_name, dimension_count
    )

    return dimension_list


def get_dimension_names(data_list, dimension_name1, num_dimensions):
    """
    Get the column headers which correspond to dimension names
    """
    dimension_names = list()
    if num_dimensions > 1:
        num_dimensions -= 1
        # Put first dimension name at start of list
        dimension_names.append(dimension_name1)
        # Add other dimension names to list
        dimension_names = dimension_names + data_list[0][:num_dimensions]
    else:
        print("single dimension")
        if dimension_name1 is None:
            dimension_name1 = "Category"
        dimension_names.append(dimension_name1)
    return dimension_names
