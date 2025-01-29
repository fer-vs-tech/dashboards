"""
Author: Kamoliddin Usmonov
Date: 2022-12-29
Description: Helper functions for ESG dashboards
"""

import logging

logger = logging.getLogger(__name__)

import datetime
import io
import os

import numpy as np
import pandas as pd
from dash import dcc
from openpyxl import load_workbook, utils

import cm_dashboards.dash_utils as dash_utils
import cm_dashboards.utilities as utilities


def get_df(handler, wvr_path, model_name="IFRS_17_PAA", replace_na=None):
    """
    Get table data from wvr
    :param handler: DB handler
    :param wvr_path: path to wvr file
    :return: DataFrame
    """
    df = handler.get_wvr_data(wvr_path, model_name)
    # Replace NaN values if provided
    if replace_na is not None:
        df = df.replace(np.nan, 0)
    return df


def prepare_dropdown_options(option_names, name="Model", default=None):
    """
    Generate dropdown options
    :param option_names: List of option names
    :return: list of dictionaries (value, label, title) for dropdown options
    """
    # logger.info("Generating dropdown options: {}".format(option_names))
    if default is not None:
        if isinstance(default, list):
            option_names = default + option_names
        else:
            option_names.insert(0, default)

    # Loop through list of option names and create a list of dropdown options
    options = []
    for option_name in option_names:
        # Flat list if it exists
        if isinstance(option_name, list):
            option_name = option_name[0]

        # Format date if flag is set
        if name == "Report date" and not isinstance(option_name, str):
            title = option_name.strftime("Report date %d-%b-%Y")
        else:
            title = f"{name} '{option_name if option_name != '' else 'All'}'"

        # Generate new option
        option = {
            "label": option_name,
            "value": option_name,
            "title": title,
        }
        options.append(option)

    # logger.info("Dropdown options for {}: {}".format(name, options))
    return options


def prepare_table_data(
    df,
    header_rows=[],
    hidden_columns=[],
    filter_column_style=None,
    additional_header=[],
    replace_zero=None,
    multi_index=False,
    show_negative_numbers=True,
):
    """
    Prepare table data
    :param df: DataFrame
    :param header_rows: list of header rows
    :param hidden_columns: list of hidden columns
    :return: tuple (data, columns, conditional_style) for dash_table.DataTable
    """
    # Check if replacement is needed
    if replace_zero is not None:
        df = df.replace({0: replace_zero})

    # Set multiindex table data, column if needed
    if multi_index:
        table_data, columns = dash_utils.set_multi_index_column_names(
            df, show_negative_numbers=show_negative_numbers
        )
    else:
        table_data = df.to_dict("records")
        table_columns = df.columns
        columns = dash_utils.set_column_names(
            table_columns,
            precision=0,
            hidden_columns=hidden_columns,
            show_negative_numbers=show_negative_numbers,
            additional_header=additional_header,
        )
    conditional_style = dash_utils.set_table_style_kics(
        columns, show_negative_numbers=show_negative_numbers
    )
    row_style = dash_utils.set_row_style(header_rows)
    style = conditional_style + row_style

    # Add conditional style for filter column
    if filter_column_style is not None:
        result = dash_utils.set_conditional_style_by_filtering(
            column=filter_column_style
        )
        style = style + result

    return table_data, columns, style


def filter_date(current_date, target_date):
    """
    Filter out dates that represent past dates than tagret date
    :param current_date: date and name of the current date (tuple)
    :return result: boolean indicating if the date should be filtered
    """
    result = current_date[0] > target_date
    return result


def pivot_df(df):
    """
    Helper function to rotate the df
    :param df: the df to rotate
    :param result: rotated result
    """
    # Rotate DF
    df = df.reset_index(drop=True).transpose()

    # Make the first row as header row
    df.columns = df.iloc[0]
    df = df[1:]

    # Convert datetime object to string in column names
    df.rename(
        columns=lambda t: t.strftime("%Y-%m-%d") if isinstance(t, datetime.date) else t,
        inplace=True,
    )

    # Reset index, rename 'index' column name
    df.reset_index(inplace=True)
    df.rename(columns={"index": ""}, inplace=True)

    return df


def calculate(value1, value2, operator):
    """
    Calculate values
    :param value1: First value
    :param value2: Second
    :param operator: Operator (str) (+, -, *, /, //)
    :return: Calculated value
    """
    # logger.info("Calculate: {} {} {}".format(value1, operator, value2))
    match operator:
        case "+":
            result = value1 + value2
        case "-":
            result = value1 - value2
        case "*":
            result = value1 * value2
        case "/":
            result = value1 / value2
        case "//":
            result = value1 // value2
        case _:
            result = value1 + value2

    return result


def get_template_df(journal_name, header=[0]):
    """
    Get DataFrame by reading template csv file
    :param journal_name: journal name (e.g. "AC")
    :return: DataFrame, list of header rows
    """
    # Get path to template csv file by journal name and make dataframe from it
    current_path = os.path.dirname(os.path.abspath(__file__))
    # go to parent directory
    current_path = os.path.dirname(current_path)

    # Get DataFrame with column
    df = pd.read_csv(
        f"{current_path}\\templates\\{journal_name}.csv",
        header=header,
        encoding="utf-8",
    )

    # Rename "Unnamed" columns to empty string
    if len(header) == 1:
        # logger.info("Renaming columns for template: {}".format(journal_name))
        df.columns = [remove_substring(col) for col in df.columns]

    return df


def remove_substring(string):
    """
    Remove substring from string
    :param string: string
    :return: string
    """
    if "." in string:
        string = string.split(".")[0]
    if "Unnamed" in string:
        string = ""
    return string


def replace_template_values(template_df, lookup_dict):
    """
    Replace template values by checking
    """
    # Use applymap to replace values if there is no formula in the template
    template_df = template_df.applymap(lambda x: apply_formatting(x, lookup_dict))
    return template_df


def apply_formatting(cell_value, lookup_dict, perform_abs=False):
    """
    Apply formatting to cell value
    :param cell_value: cell value
    :param lookup_dict: lookup dictionary
    :return: formatted cell value with replacement value if applicable
    """
    if (
        pd.isna(cell_value)
        or pd.isnull(cell_value)
        or cell_value == ""
        or isinstance(cell_value, int)
    ):
        if str(cell_value).lower() == "true":
            return str(cell_value).upper()
        return cell_value

    # Remove substring from cell value if possible
    try:
        new_value = lookup_dict.get(cell_value, cell_value)
    except Exception as e:
        logger.info("Error occurred: {}".format(e))
        new_value = cell_value

    # Cast cell value to float if possible
    if not isinstance(new_value, (int, float, datetime.date)) and new_value.isdigit():
        new_value = float(new_value)

    # logger.info("Cell value: {}, new value: {}".format(cell_value, new_value))
    return new_value if new_value != 0 else "-"


def generate_report_file(results, filename, group_id):
    """
    Write data to the template excel file (openpyxl engine)
    :param results: Results (dict)
    :param filename: Filename to save the file as
    :param group_id: Group ID that used to rewrite sheet name
    :return updated excel file (in memory buffer)
    """
    # If results is a list, get the first element as lookup source
    if isinstance(results, list):
        results = results[0]

    # Define report type
    if group_id in ["PAA", "PAA_Reins"]:
        report_types = {
            "PAA": "PAA_Output",
            "PAA_Reins": "PAA_Output_Reins",
        }
        report_type = report_types.get(group_id, "PAA_Output")
        template_file = "companylevel_template.xlsx"
    else:
        report_type = "PAA_Temp"
        template_file = "grouplevel_template.xlsx"

    # File path to the template
    current_path = os.path.dirname(os.path.abspath(__file__))
    current_path = os.path.dirname(current_path)
    template_file = os.path.join(current_path, "templates", template_file)
    # Load the template file
    book = load_workbook(template_file, data_only=True)
    book.calculation_on_load = True
    # Set the first sheet as the active sheet
    book.active = 0
    # Set the active cell to the top-left cell of the worksheet
    book.active.cell(row=1, column=1)

    # Loop through each sheet that contains its data and dashboard ids
    try:
        for dashboard_id in dashboards_list().keys():
            # Define named range destination
            current_range = dashboard_id
            current_range = current_range.replace("-", "_")
            try:
                coords, sheet = utilities.get_range_destination(book, current_range)
            except Exception:
                # Skip if range is not found
                continue

            # Get numerical dimentions of coordinates
            boundaries = utils.cell.range_boundaries(coords)
            first_col = boundaries[0]
            first_row = boundaries[1]
            last_col = boundaries[2]
            last_row = boundaries[3]

            # Loop over each row in the range
            for row in sheet.iter_rows(
                min_row=first_row,
                max_row=last_row,
                min_col=first_col,
                max_col=last_col,
            ):
                # Loop over each cell in the row
                for cell in row:
                    if cell.value is None or cell.value == "":
                        continue
                    # logger.info(f"Current cell value: {cell.value}")

                    # Replace the value if it is found in results
                    if cell.value in results:
                        # logger.info(
                        #     f"Found match: {cell.value} - {results[cell.value]}"
                        # )
                        cell.value = results[cell.value]
                        # Apply custom number format
                        cell.number_format = (
                            '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
                        )

    except Exception as error:
        logger.info(f"Error occurred while processing report file: {error}")
        raise utilities.add_exception_info(error, "Error while processing report file")

    # Rename the sheet
    book.active.title = report_type

    # Save the changes to the memory buffer and close the workbook
    in_mem_file = io.BytesIO()
    book.save(in_mem_file)
    book.close()

    # Reset the pointer to the beginning of the file
    in_mem_file.seek(0)

    return dcc.send_bytes(in_mem_file.getvalue(), filename)


def dashboards_list():
    """
    Return list of dashboards
    """
    return {
        "dashboard_1": {
            "header": [0, 1],
            "header_rows": [0, 11, 12, 13, 14, 19],
            "label": "未到期责任负债及已发生赔款负债余额调节表",
        },
        "dashboard_2": {
            "header": [0],
            "header_rows": [0, 3, 4, 7, 11],
            "label": "利润表",
        },
        "dashboard_3": {
            "header": [0],
            "header_rows": [0, 4, 9],
            "label": "其他综合收益",
        },
        "dashboard_4": {
            "header": [0],
            "header_rows": [0, 1, 5],
            "label": "资产负债表",
        },
    }
